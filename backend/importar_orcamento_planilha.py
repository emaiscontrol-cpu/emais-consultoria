import os
import openpyxl
from sqlalchemy import text
from database import engine, SessionLocal, Base
import models

def importar():
    # 1. Certifica que a tabela nova existe
    print("Verificando/Criando tabelas no banco de dados...")
    Base.metadata.create_all(bind=engine)
    
    # 2. Localiza o arquivo de planilha
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "..", "ORCAMENTO", "FC - 2025 - ORÇAMENTO_A1.xlsx")
    
    if not os.path.exists(file_path):
        print(f"Erro: Arquivo não encontrado em {file_path}")
        return
        
    print(f"Carregando planilha: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if "CLAUDE" not in wb.sheetnames:
        print("Erro: Aba 'CLAUDE' não encontrada no arquivo!")
        return
        
    sheet = wb["CLAUDE"]
    db = SessionLocal()
    
    # Parâmetros
    CLIENTE_ID = 10  # Rio das Pedras
    ANO = 2026
    VERSAO = "Original"
    
    try:
        # 3. Limpa orçamentos anteriores para evitar duplicados
        print(f"Limpando orçamento antigo de {ANO} para o cliente ID {CLIENTE_ID}...")
        db.execute(
            text("DELETE FROM fc_orcamento WHERE cliente_id = :cid AND ano = :ano AND versao = :ver"),
            {"cid": CLIENTE_ID, "ano": ANO, "ver": VERSAO}
        )
        db.commit()
        
        # Load template lines to map row index (ordem) to DB slug
        template_map = {}
        db_lines = db.execute(text("""
            SELECT tl.ordem, tl.agrupamento_slug
            FROM ref_template_linhas tl
            JOIN ref_templates t ON t.id = tl.template_id
            WHERE t.tipo = 'fluxo_caixa' AND t.ativo = true
        """)).fetchall()
        for line in db_lines:
            if line.agrupamento_slug:
                template_map[line.ordem] = line.agrupamento_slug

        # 4. Processa as linhas
        registros_inseridos = 0
        total_valor = 0.0
        
        # O cabeçalho dos meses está na linha 2. Os dados começam na linha 4
        rows = list(sheet.iter_rows(values_only=True))
        
        for r_idx, row in enumerate(rows, start=1):
            if r_idx < 4:
                continue
                
            db_slug = template_map.get(r_idx)
            if not db_slug:
                slug = row[0]
                if not slug or not str(slug).strip():
                    continue
                db_slug = str(slug).strip()
            
            # Para cada mês de 1 a 12
            for m in range(1, 13):
                col_idx = 5 + (m - 1) * 5
                if col_idx >= len(row):
                    break
                    
                val_raw = row[col_idx]
                
                # Conversão segura para float
                try:
                    val = float(val_raw) if val_raw is not None else 0.0
                except (ValueError, TypeError):
                    val = 0.0
                    
                # Cria o registro do orçamento
                orcamento_row = models.FCOrcamento(
                    cliente_id=CLIENTE_ID,
                    agrupamento_slug=db_slug,
                    ano=ANO,
                    mes=m,
                    valor=val,
                    versao=VERSAO
                )
                db.add(orcamento_row)
                registros_inseridos += 1
                total_valor += val
                
        db.commit()
        print(f"Importação concluída com sucesso!")
        print(f"  - Total de registros inseridos: {registros_inseridos}")
        print(f"  - Valor total orçado acumulado: R$ {total_valor:,.2f}")
        
    except Exception as e:
        db.rollback()
        print(f"Erro durante a importação: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    importar()

"""
Seed DRE — Shopping
Lê ORCAMENTO/DRE - Shopping.xlsx e insere os itens no plano "Varejo SC — Padrão"
com modulo="D" (ou "D,O" para contas que coincidem com o Orçamento).
Regras:
  - N1 e N2 → tipo="TT", conta=código numérico (ex: "3000")
  - N3       → tipo="NN", conta=código contábil ERP (ex: "3.01.001.0002.00001")
  - REMUNERAÇÃO (3083-3089) sem código ERP → atribuídos 4.01.003.0011.00001..00007
  - modulo D,O para contas que coincidem com o Orçamento, D para as demais
  - movimento detectado por palavras-chave na descrição
Idempotente: verifica por conta antes de inserir.
"""
import re
from pathlib import Path
from models import Plano, PlanoItem

# Códigos DRE que também existem no Orçamento → modulo D,O
MATCH_DO = {
    3000, 3030, 3050, 3051, 3060, 3061, 3080, 3081, 3120, 3121,
    3122, 3135, 3150, 3165, 3175, 3195, 3205, 3330, 3335, 3340,
    3361, 3380, 3500, 3390, 3420,
}

# Contas manuais para itens de REMUNERAÇÃO sem código ERP no arquivo
REMUNERACAO = {
    3083: "4.01.003.0011.00001",
    3084: "4.01.003.0011.00002",
    3085: "4.01.003.0011.00003",
    3086: "4.01.003.0011.00004",
    3087: "4.01.003.0011.00005",
    3088: "4.01.003.0011.00006",
    3089: "4.01.003.0011.00007",
}

def _agrupamento(cod: int) -> str:
    if cod <= 3029: return "RECEITA"
    if cod <= 3049: return "DEDUCOES"
    if cod == 3050: return "TOTAL_RECEITA"
    if cod <= 3059: return "CUSTOS_VAR"
    if cod == 3060: return "MARGEM_VENDA"
    if cod <= 3079: return "DESP_VAR"
    if cod == 3080: return "MARGEM_CONTR"
    if cod <= 3119: return "PESSOAL"
    if cod == 3120: return "MARGEM_CONTR2"
    if cod <= 3389: return "CUSTOS_FIX_IND"
    if cod <= 3419: return "OUTRAS_REC"
    if cod <= 3429: return "EBITDA"
    if cod <= 3469: return "RESULTADO_FIN"
    return "RESULTADO"

def _movimento(nivel: int, desc: str):
    if nivel in (1, 2):
        return None
    d = desc.upper()
    entradas = ["RECEITA", "VENDA", "BENEFICIO", "BENEFÍCIO",
                "JUROS RECEB", "RENDIMENTO APLIC", "DESCONTO OBTID",
                "OUTRAS RECEIT", "RECEITAS COM ACORDOS"]
    if any(k in d for k in entradas):
        return "Entrada"
    return "Saída"


def seed_dre(db):
    plano = db.query(Plano).filter(Plano.nome == "Varejo SC — Padrão").first()
    if not plano:
        print("[seed_dre] Plano 'Varejo SC — Padrão' não encontrado. Execute seed_orcamento primeiro.")
        return

    # Contas já existentes no plano (evita duplicatas)
    existentes = {i.conta for i in plano.itens if i.conta}

    xlsx = Path(__file__).parent.parent / "ORCAMENTO" / "DRE - Shopping.xlsx"
    if not xlsx.exists():
        print(f"[seed_dre] Arquivo não encontrado: {xlsx}")
        return

    try:
        import openpyxl
    except ImportError:
        print("[seed_dre] openpyxl não instalado — seed DRE ignorado.")
        return

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ws = wb["Consolidado Goias DF"]

    novos = 0
    ordem_max = max((i.ordem for i in plano.itens), default=0)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        c1, c2, c3, _, cod_raw = row[0], row[1], row[2], row[3], row[4]
        if cod_raw is None:
            continue
        try:
            cod = int(cod_raw)
        except (ValueError, TypeError):
            continue

        if   c1: nivel, nome_raw = 1, str(c1).strip()
        elif c2: nivel, nome_raw = 2, str(c2).strip()
        elif c3: nivel, nome_raw = 3, str(c3).strip()
        else:    continue

        # Extrai código contábil ERP embutido na descrição (N3)
        conta_erp = None
        desc = nome_raw
        # Padrão "X.XX.XXX.XXXX.XXXXX - Descrição"
        m = re.match(r'^(\d+\.\d+\.\d+\.\d+\.\d+)\s*[-–]\s*(.+)$', nome_raw)
        if m:
            conta_erp, desc = m.group(1), m.group(2).strip()
        else:
            # Padrão "X.XX.XXX.XXXX.XXXXX  Descrição" (separado por espaço)
            m2 = re.match(r'^(\d+\.\d+\.\d+\.\d+\.\d+)\s+(.+)$', nome_raw)
            if m2:
                conta_erp, desc = m2.group(1), m2.group(2).strip()

        # Contas manuais para REMUNERAÇÃO
        if cod in REMUNERACAO and not conta_erp:
            conta_erp = REMUNERACAO[cod]

        # Chave de identificação: ERP code para N3, código numérico para N1/N2
        if nivel in (1, 2):
            tipo        = "TT"
            conta_final = str(cod)     # ex: "3000"
        else:
            tipo        = "NN"
            conta_final = conta_erp or str(cod)

        # Pula se já existe
        if conta_final in existentes:
            continue

        ordem_max += 1
        modulo    = "D,O" if cod in MATCH_DO else "D"
        movimento = _movimento(nivel, desc)
        agrup     = _agrupamento(cod)

        db.add(PlanoItem(
            plano_id    = plano.id,
            ordem       = ordem_max,
            agrupamento = agrup,
            conta       = conta_final,
            descricao   = desc,
            tipo        = tipo,
            modulo      = modulo,
            movimento   = movimento,
        ))
        existentes.add(conta_final)
        novos += 1

    if novos > 0:
        db.commit()
        print(f"[seed_dre] {novos} itens DRE adicionados ao plano '{plano.nome}'.")
    else:
        print("[seed_dre] Nenhum item novo para adicionar (já importados).")

    _atualizar_formulas_dre(db, plano)


# Fórmulas dos TTs resultado — baseadas na estrutura do Excel SC 3.11
# Chave = agrupamento do item TT sem filhos NN diretos
_FORMULAS_DRE = {
    "TOTAL_RECEITA":  "RECEITA - DEDUCOES",
    "MARGEM_VENDA":   "TOTAL_RECEITA - CUSTOS_VAR",
    "MARGEM_CONTR":   "MARGEM_VENDA - DESP_VAR",
    "MARGEM_CONTR2":  "MARGEM_CONTR - PESSOAL",
    "EBITDA":         "MARGEM_CONTR2 - CUSTOS_FIX_IND + OUTRAS_REC",
    "RESULTADO":      "EBITDA - RESULTADO_FIN",
}


def _atualizar_formulas_dre(db, plano):
    """Atualiza o campo formula nos TTs resultado do DRE Shopping."""
    alterados = 0
    for item in plano.itens:
        nova = _FORMULAS_DRE.get(item.agrupamento)
        if nova and item.formula != nova:
            item.formula = nova
            alterados += 1
    if alterados:
        db.commit()
        print(f"[seed_dre] {alterados} fórmulas de resultado atualizadas.")

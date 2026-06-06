"""
Importação do DRE Shopping — dados históricos 2025.
Lê as 12 abas "Xxx Ant" do arquivo DRE - Shopping.xlsx e popula
a tabela orcamento_unidade_valores para o cliente Shopping Cosméticos.

Uso:
    python import_dre_shopping.py

Pode ser executado múltiplas vezes (idempotente via UPSERT).
"""

import sys
import os

# Garante que o diretório do backend está no path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

from database import SessionLocal
from sqlalchemy import text

# ── Configuração ──────────────────────────────────────────────────────────────

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "ORCAMENTO", "DRE - Shopping.xlsx"
)

CLIENTE_NOME = "Shopping Cosmeticos"
ANO = 2025

# Mapeamento aba → número do mês (notar "Abr Ant " com espaço)
MESES_ABAS = {
    1:  "Jan Ant",
    2:  "Fev Ant",
    3:  "Mar Ant",
    4:  "Abr Ant ",
    5:  "Mai Ant",
    6:  "Jun Ant",
    7:  "Jul Ant",
    8:  "Ago Ant",
    9:  "Set Ant",
    10: "Out Ant",
    11: "Nov Ant",
    12: "Dez Ant",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _extrair_conta(cell_value) -> str:
    """
    Extrai o código de conta de um valor de célula.
    Exemplos:
      "3.01.001.0008.00001 - Receitas Com Vendas Em Dinheiro" → "3.01.001.0008.00001"
      "3001" → "3001"
    """
    if cell_value is None:
        return ""
    s = str(cell_value).strip()
    if " - " in s:
        s = s.split(" - ")[0].strip()
    return s


def _is_conta_n3(v) -> bool:
    """Retorna True se o valor parece um código de conta analítica (N3)."""
    if not v:
        return False
    s = str(v).strip()
    # Código analítico: formato x.xx.xxx.xxxx.xxxxx
    return "." in s and len(s.split(".")) >= 4


def _is_dre_code(v) -> bool:
    """Retorna True se parece um código DRE numérico como 3001, 3002, ..."""
    if not v:
        return False
    try:
        n = int(str(v).strip())
        return 3000 <= n <= 4000
    except ValueError:
        return False


# ── Importação ────────────────────────────────────────────────────────────────

def importar(excel_path: str = EXCEL_PATH):
    db = SessionLocal()

    try:
        # Garantir que a tabela existe (caso rode antes do restart do serviço)
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS orcamento_unidade_valores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plano_item_id INTEGER NOT NULL REFERENCES planos_itens(id),
                cliente_id INTEGER NOT NULL REFERENCES clientes(id),
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                unidade TEXT NOT NULL,
                valor REAL DEFAULT 0.0,
                UNIQUE(plano_item_id, cliente_id, ano, mes, unidade)
            )
        """))
        db.commit()

        # Buscar cliente
        result = db.execute(
            text("SELECT id FROM clientes WHERE razao_social = :nome"),
            {"nome": CLIENTE_NOME}
        ).fetchone()
        if not result:
            print(f"[ERRO] Cliente '{CLIENTE_NOME}' não encontrado no banco.")
            print("       Cadastre o cliente primeiro via sistema.")
            return
        cliente_id = result[0]
        print(f"Cliente: {CLIENTE_NOME} (id={cliente_id})")

        # Buscar plano vinculado
        result = db.execute(
            text("SELECT plano_id FROM cliente_plano WHERE cliente_id = :cid"),
            {"cid": cliente_id}
        ).fetchone()
        if not result:
            print(f"[ERRO] Cliente não possui plano vinculado.")
            return
        plano_id = result[0]
        print(f"Plano vinculado: id={plano_id}")

        # Carregar mapa de contas do plano → {conta: plano_item_id}
        rows = db.execute(
            text("SELECT id, conta FROM planos_itens WHERE plano_id = :pid AND conta IS NOT NULL"),
            {"pid": plano_id}
        ).fetchall()
        conta_map = {str(r[1]).strip(): r[0] for r in rows}
        print(f"Mapa de contas carregado: {len(conta_map)} itens")

        # Apagar dados anteriores do cliente/ano (re-import limpo)
        deleted = db.execute(
            text("DELETE FROM orcamento_unidade_valores WHERE cliente_id=:cid AND ano=:ano"),
            {"cid": cliente_id, "ano": ANO}
        ).rowcount
        if deleted:
            print(f"Removidos {deleted} registros anteriores do ano {ANO}")
        db.commit()

        # Abrir Excel
        print(f"\nAbrindo Excel: {excel_path}")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        total_inseridos = 0
        total_nao_mapeados = set()

        for mes, aba_nome in MESES_ABAS.items():
            if aba_nome not in wb.sheetnames:
                # Tentar sem/com espaços
                candidatos = [n for n in wb.sheetnames if n.strip() == aba_nome.strip()]
                if candidatos:
                    aba_nome = candidatos[0]
                else:
                    print(f"  [AVISO] Aba '{aba_nome}' não encontrada, pulando mês {mes}")
                    continue

            ws = wb[aba_nome]
            all_rows = list(ws.iter_rows(values_only=True))
            print(f"  Mês {mes:2d} ({aba_nome.strip()}): {len(all_rows)} linhas")

            # Linha 6 (índice 5): códigos das unidades (colunas ímpares de 7 em diante)
            r6 = all_rows[5] if len(all_rows) > 5 else []
            unidade_cols = []  # [(col_index, codigo_unidade), ...]
            for i, v in enumerate(r6):
                if v is not None and str(v).strip() and str(v).strip() != '%':
                    # Apenas colunas de valor (não % — que são as seguintes ao par)
                    unidade_cols.append((i, str(v).strip()))

            # Linhas de dados (a partir da linha 8, índice 7)
            inseridos_mes = 0
            for row in all_rows[7:]:
                c1 = row[0]   # Nome N1
                c2 = row[1]   # Nome N2
                c3 = row[2]   # Código N3 analítico
                c5 = row[4]   # Código DRE (TT)

                conta = ""
                if _is_conta_n3(c3):
                    conta = _extrair_conta(c3)
                elif _is_dre_code(c5):
                    conta = str(int(float(str(c5).strip())))
                else:
                    continue

                plano_item_id = conta_map.get(conta)
                if not plano_item_id:
                    total_nao_mapeados.add(conta)
                    continue

                # Para cada unidade, inserir valor
                for col_idx, unidade_cod in unidade_cols:
                    if col_idx >= len(row):
                        continue
                    raw = row[col_idx]
                    if raw is None or str(raw).strip() in ('', '#DIV/0!', '#REF!', '#VALUE!', '#N/A'):
                        continue
                    try:
                        valor = float(raw)
                    except (TypeError, ValueError):
                        continue

                    db.execute(
                        text("""
                            INSERT INTO orcamento_unidade_valores
                                (plano_item_id, cliente_id, ano, mes, unidade, valor)
                            VALUES (:iid, :cid, :ano, :mes, :uni, :val)
                            ON CONFLICT(plano_item_id, cliente_id, ano, mes, unidade)
                            DO UPDATE SET valor=excluded.valor
                        """),
                        {
                            "iid": plano_item_id,
                            "cid": cliente_id,
                            "ano": ANO,
                            "mes": mes,
                            "uni": unidade_cod,
                            "val": valor,
                        }
                    )
                    inseridos_mes += 1

            db.commit()
            total_inseridos += inseridos_mes
            print(f"    -> {inseridos_mes} valores inseridos")

        wb.close()

        print(f"\n[OK] Importacao concluida: {total_inseridos} registros totais")
        if total_nao_mapeados:
            print(f"[AVISO] {len(total_nao_mapeados)} contas sem mapeamento no plano:")
            for c in sorted(total_nao_mapeados)[:20]:
                print(f"   {c}")

    finally:
        db.close()


if __name__ == "__main__":
    importar()

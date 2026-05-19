from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from datetime import datetime, timezone, timedelta
from io import BytesIO
from database import get_db
from auth import get_usuario_atual
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import models

router = APIRouter()

STATUS_PT = {
    "planejamento": "Planejamento",
    "em_andamento": "Em andamento",
    "pausado":      "Pausado",
    "concluido":    "Concluído",
    "atrasado":     "Atrasado",
    "pendente":     "Pendente",
    "bloqueada":    "Bloqueada",
    "concluida":    "Concluída",
    "atrasada":     "Atrasada",
    "aguard_validacao": "Aguard. validação",
}

HDR_FILL  = PatternFill("solid", fgColor="1E40AF")
HDR_FONT  = Font(bold=True, color="FFFFFF")
ALT_FILL  = PatternFill("solid", fgColor="EFF6FF")


def _fmt_data(dt):
    if not dt:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def _cabecalho(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
        letra = c.column_letter
        ws.column_dimensions[letra].width = max(18, len(h) + 4)


@router.get("/projetos/excel")
def relatorio_projetos(db: Session = Depends(get_db), usuario=Depends(get_usuario_atual)):
    wb = Workbook()

    # ── Aba 1: Projetos ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Projetos"
    _cabecalho(ws1, ["Projeto", "Cliente", "Status", "Progresso %", "Início", "Prev. Fim", "Fases", "Concluídas"])

    q = db.query(models.Projeto).options(
        joinedload(models.Projeto.cliente),
        joinedload(models.Projeto.fases),
    )
    if usuario.perfil == "cliente":
        q = q.filter(models.Projeto.cliente_id == usuario.cliente_id)
    projetos = q.all()

    for i, p in enumerate(projetos, 2):
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [
            p.nome,
            p.cliente.razao_social if p.cliente else "",
            STATUS_PT.get(p.status.value, p.status.value),
            p.progresso,
            _fmt_data(p.data_inicio),
            _fmt_data(p.data_fim_prev),
            len(p.fases),
            sum(1 for f in p.fases if f.status.value == "concluida"),
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=i, column=col, value=val)
            if fill:
                c.fill = fill

    # ── Aba 2: Tarefas ───────────────────────────────────
    ws2 = wb.create_sheet("Tarefas")
    _cabecalho(ws2, ["Projeto", "Fase", "Tarefa", "Status", "Responsável", "Prazo", "% Concluído", "Req. Validação"])

    row = 2
    for p in projetos:
        for fase in p.fases:
            for t in fase.tarefas:
                resp = t.responsavel.nome if t.responsavel else ""
                fill = ALT_FILL if row % 2 == 0 else None
                vals = [
                    p.nome, fase.nome, t.nome,
                    STATUS_PT.get(t.status.value, t.status.value),
                    resp, _fmt_data(t.data_prazo), t.percentual,
                    "Sim" if t.requer_validacao else "Não",
                ]
                for col, val in enumerate(vals, 1):
                    c = ws2.cell(row=row, column=col, value=val)
                    if fill:
                        c.fill = fill
                row += 1

    # ── Aba 3: Histórico ─────────────────────────────────
    ws3 = wb.create_sheet("Histórico")
    _cabecalho(ws3, ["Data", "Usuário", "Ação", "Descrição", "Projeto"])

    logs = db.query(models.LogAtividade).order_by(
        models.LogAtividade.criado_em.desc()
    ).limit(500).all()

    for i, lg in enumerate(logs, 2):
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [
            _fmt_data(lg.criado_em),
            lg.usuario.nome if lg.usuario else "",
            lg.acao,
            lg.descricao,
            lg.projeto.nome if lg.projeto else "",
        ]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=i, column=col, value=val)
            if fill:
                c.fill = fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome = f"relatorio_emais_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"},
    )


def _val(enum_or_str):
    return enum_or_str.value if hasattr(enum_or_str, "value") else str(enum_or_str)

def _fmt(dt):
    if not dt: return None
    return dt.strftime("%Y-%m-%d")


@router.get("/graficos/{projeto_id}")
def graficos(projeto_id: int, db: Session = Depends(get_db), usuario=Depends(get_usuario_atual)):
    projeto = db.query(models.Projeto).options(
        joinedload(models.Projeto.fases).joinedload(models.Fase.tarefas)
    ).get(projeto_id)
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")

    hoje = datetime.now(timezone.utc).date()

    # ── Coleta todas as tarefas ativas ───────────────────
    todas = []
    for f in projeto.fases:
        for t in f.tarefas:
            if getattr(t, "ativo", True):
                todas.append((f, t))

    def _status_cat(status_val):
        if status_val == "concluida":   return "concluida"
        if status_val in ("em_andamento", "aguard_validacao", "aguard_valid"): return "em_andamento"
        return "parada"

    # ── 1. Pizza geral ───────────────────────────────────
    cont = {"concluida": 0, "em_andamento": 0, "parada": 0}
    for _, t in todas:
        cont[_status_cat(_val(t.status))] += 1

    pizza_geral = [
        {"name": "Feito",        "value": cont["concluida"],    "color": "#3B6D11"},
        {"name": "Em andamento", "value": cont["em_andamento"], "color": "#C97D10"},
        {"name": "Parado",       "value": cont["parada"],       "color": "#A32D2D"},
    ]

    # ── 2. Pizza por fase ────────────────────────────────
    por_fase = []
    for fase in projeto.fases:
        ativas = [t for t in fase.tarefas if getattr(t, "ativo", True)]
        c = sum(1 for t in ativas if _val(t.status) == "concluida")
        a = sum(1 for t in ativas if _val(t.status) in ("em_andamento","aguard_validacao","aguard_valid"))
        p = max(0, len(ativas) - c - a)
        por_fase.append({
            "nome": fase.nome,
            "ordem": fase.ordem,
            "progresso": fase.progresso,
            "status": _val(fase.status),
            "total": len(ativas),
            "concluidas": c,
            "em_andamento": a,
            "paradas": p,
            "pizza": [
                {"name": "Feito",        "value": c, "color": "#3B6D11"},
                {"name": "Em andamento", "value": a, "color": "#C97D10"},
                {"name": "Parado",       "value": p, "color": "#A32D2D"},
            ]
        })

    # ── 3. Volume por usuário ────────────────────────────
    umap = {}
    for _, t in todas:
        uid = t.responsavel_id
        nome = t.responsavel.nome if t.responsavel else None
        if not uid or not nome:
            continue
        if uid not in umap:
            umap[uid] = {"nome": nome, "tarefas": []}
        umap[uid]["tarefas"].append({
            "nome": t.nome,
            "status": _val(t.status),
            "cat": _status_cat(_val(t.status)),
        })

    usuarios = []
    for u in umap.values():
        total = len(u["tarefas"])
        conc  = sum(1 for t in u["tarefas"] if t["cat"] == "concluida")
        perc  = round(conc / total * 100) if total else 0
        usuarios.append({
            "nome": u["nome"],
            "total": total,
            "concluidas": conc,
            "percentual": perc,
            "tarefas": u["tarefas"],
        })
    usuarios.sort(key=lambda x: -x["percentual"])

    # ── 4. Burndown ──────────────────────────────────────
    inicio = projeto.data_inicio.date() if projeto.data_inicio else hoje - timedelta(days=30)
    fim    = projeto.data_fim_prev.date() if projeto.data_fim_prev else hoje + timedelta(days=30)
    if fim < hoje:
        fim = hoje

    total_t = len(todas)
    conclusoes = sorted([t.data_conclusao.date() for _, t in todas if t.data_conclusao])

    duracao = max(1, (fim - inicio).days)
    step    = max(1, duracao // 18)
    burndown = []
    d = inicio
    while d <= fim:
        dias_passados = (d - inicio).days
        ideal = round(total_t * max(0, 1 - dias_passados / duracao), 1)
        real  = total_t - sum(1 for c in conclusoes if c <= d)
        burndown.append({
            "data":  d.strftime("%d/%m"),
            "ideal": ideal,
            "real":  real if d <= hoje else None,
        })
        d += timedelta(days=step)

    # ── 5. Gantt ─────────────────────────────────────────
    gantt = []
    for fase in projeto.fases:
        fi = fase.data_inicio.date()   if fase.data_inicio   else None
        ff = fase.data_fim_prev.date() if fase.data_fim_prev else None
        gantt.append({
            "id":       f"f{fase.id}",
            "nome":     fase.nome,
            "tipo":     "fase",
            "inicio":   _fmt(fase.data_inicio),
            "fim":      _fmt(fase.data_fim_prev),
            "progresso": fase.progresso,
            "status":   _val(fase.status),
        })
        for t in fase.tarefas:
            if not getattr(t, "ativo", True): continue
            gantt.append({
                "id":       f"t{t.id}",
                "nome":     t.nome,
                "tipo":     "tarefa",
                "fase":     fase.nome,
                "inicio":   _fmt(t.data_inicio),
                "fim":      _fmt(t.data_prazo),
                "progresso": t.percentual,
                "status":   _val(t.status),
            })

    return {
        "projeto":      {"id": projeto.id, "nome": projeto.nome, "progresso": projeto.progresso,
                         "status": _val(projeto.status),
                         "data_inicio": _fmt(projeto.data_inicio),
                         "data_fim_prev": _fmt(projeto.data_fim_prev)},
        "pizza_geral":  pizza_geral,
        "por_fase":     por_fase,
        "usuarios":     usuarios,
        "burndown":     burndown,
        "gantt":        gantt,
        "total_tarefas": total_t,
    }

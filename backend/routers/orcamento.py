from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, List
from collections import defaultdict
import openpyxl
import io
import os
import re
import json
import httpx
from pydantic import BaseModel
import models
from database import get_db
from auth import get_usuario_atual, requer_perfil, verificar_tenant
from routers.fc_exec import _parse_compound_slug, _eval_formula

router = APIRouter()

@router.get("/clientes")
def listar_clientes_com_analise_gerencial(
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual)
):
    """Lista apenas clientes ativos que possuem o módulo de Análises Gerenciais habilitado."""
    return db.query(models.Cliente).filter(
        models.Cliente.ativo == True,
        models.Cliente.modulo_analises_gerenciais == True
    ).all()

@router.get("/cliente/{cliente_id}/ano/{ano}")
def obter_orcamento(
    cliente_id: int,
    ano: int,
    versao: str = "Original",
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual)
):
    """Retorna os valores orçados por agrupamento_slug para um ano e versão."""
    orcamentos = db.query(models.FCOrcamento).filter(
        models.FCOrcamento.cliente_id == cliente_id,
        models.FCOrcamento.ano == ano,
        models.FCOrcamento.versao == versao
    ).all()
    
    data = {}
    for o in orcamentos:
        if o.agrupamento_slug not in data:
            data[o.agrupamento_slug] = [0.0] * 12
        if 1 <= o.mes <= 12:
            data[o.agrupamento_slug][o.mes - 1] = o.valor
            
    return [
        {
            "agrupamento_slug": slug,
            "versao": versao,
            "valores_mensais": vals
        }
        for slug, vals in data.items()
    ]


@router.post("/cliente/{cliente_id}/ano/{ano}/importar")
def importar_orcamento(
    cliente_id: int,
    ano: int,
    file: UploadFile = File(...),
    versao: str = "Original",
    db: Session = Depends(get_db),
    usuario=Depends(requer_perfil("admin", "consultor"))
):
    """Lê planilha Excel (aba CLAUDE) e salva os valores de orçamento no banco."""
    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler arquivo Excel: {str(e)}")
        
    if "CLAUDE" not in wb.sheetnames:
        raise HTTPException(400, "Aba 'CLAUDE' não encontrada no arquivo.")
        
    sheet = wb["CLAUDE"]
    
    # Load template lines to map row index (ordem) to DB slug
    template_map = {}
    try:
        db_lines = db.execute(text("""
            SELECT tl.ordem, tl.agrupamento_slug
            FROM ref_template_linhas tl
            JOIN ref_templates t ON t.id = tl.template_id
            WHERE t.tipo = 'fluxo_caixa' AND t.ativo = true
        """)).fetchall()
        for line in db_lines:
            if line.agrupamento_slug:
                template_map[line.ordem] = line.agrupamento_slug
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler templates do banco: {str(e)}")
            
    novos_registros = []
    rows = list(sheet.iter_rows(values_only=True))
    
    try:
        for r_idx, row in enumerate(rows, start=1):
            if r_idx < 4:
                continue
                
            db_slug = template_map.get(r_idx)
            if not db_slug:
                slug = row[0]
                if not slug or not str(slug).strip():
                    continue
                db_slug = str(slug).strip()
                
            for m in range(1, 13):
                col_idx = 5 + (m - 1) * 5
                if col_idx >= len(row):
                    break
                    
                val_raw = row[col_idx]
                try:
                    val = float(val_raw) if val_raw is not None else 0.0
                except (ValueError, TypeError):
                    val = 0.0
                    
                novos_registros.append(models.FCOrcamento(
                    cliente_id=cliente_id,
                    agrupamento_slug=db_slug,
                    ano=ano,
                    mes=m,
                    valor=val,
                    versao=versao
                ))
    except Exception as e:
        raise HTTPException(400, f"Erro ao processar dados da planilha: {str(e)}")

    # Fase de gravação transacional única
    try:
        # Remove registros anteriores da mesma versão
        db.execute(
            text("DELETE FROM fc_orcamento WHERE cliente_id = :cid AND ano = :ano AND versao = :ver"),
            {"cid": cliente_id, "ano": ano, "ver": versao}
        )
        
        # Insere novos registros
        for reg in novos_registros:
            db.add(reg)
            
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(400, f"Erro ao salvar orçamento no banco de dados: {str(e)}")
        
    return {"success": True, "registros_inseridos": len(novos_registros)}



@router.get("/cliente/{cliente_id}/comparativo")
def obter_comparativo(
    cliente_id: int,
    ano: int,
    versao: str = "Original",
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual)
):
    """Calcula comparativo Realizado vs Orçado lado a lado por mês."""
    # 1. Carrega as linhas de template do fluxo de caixa
    rows = db.execute(text("""
        SELECT tl.ordem, tl.rotulo, tl.tipo, tl.negrito_totalizador,
               tl.agrupamento_slug, tl.formula_texto
        FROM ref_template_linhas tl
        JOIN ref_templates t ON t.id = tl.template_id
        WHERE t.tipo = 'fluxo_caixa' AND t.ativo = true
        ORDER BY tl.ordem
    """)).fetchall()
    
    if not rows:
        return []
        
    # 2. Busca Realizado agrupado por slug e mês
    fc_raw = db.execute(text(
        "SELECT LOWER(agrupamento_slug) as slug, mes, COALESCE(SUM(valor),0) as total "
        "FROM fc_lancamentos WHERE cliente_id=:cid AND ano=:ano "
        "GROUP BY LOWER(agrupamento_slug), mes"
    ), {"cid": cliente_id, "ano": ano}).fetchall()
    by_mes_real: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    for r in fc_raw:
        by_mes_real[r.slug][r.mes] += float(r.total)
        
    # 3. Busca Orçado agrupado por slug e mês
    orc_raw = db.execute(text(
        "SELECT LOWER(agrupamento_slug) as slug, mes, COALESCE(SUM(valor),0) as total "
        "FROM fc_orcamento WHERE cliente_id=:cid AND ano=:ano AND versao=:ver "
        "GROUP BY LOWER(agrupamento_slug), mes"
    ), {"cid": cliente_id, "ano": ano, "ver": versao}).fetchall()
    by_mes_orc: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    for r in orc_raw:
        by_mes_orc[r.slug][r.mes] += float(r.total)
        
    # Helpers de cálculo de chaves compostas
    def get_val_real(slug_str: Optional[str]) -> dict[int, float]:
        components = _parse_compound_slug(slug_str)
        res = defaultdict(float)
        seen = set()
        for slug_comp, sign in components:
            key = slug_comp.lower()
            if key in seen:
                continue
            seen.add(key)
            for m in range(1, 13):
                res[m] += sign * by_mes_real[key].get(m, 0.0)
        return res
        
    def get_val_orc(slug_str: Optional[str]) -> dict[int, float]:
        if not slug_str:
            return defaultdict(float)
        full_key = slug_str.lower()
        if full_key in by_mes_orc:
            res = defaultdict(float)
            for m in range(1, 13):
                res[m] = by_mes_orc[full_key].get(m, 0.0)
            return res
            
        components = _parse_compound_slug(slug_str)
        res = defaultdict(float)
        seen = set()
        for slug_comp, sign in components:
            key = slug_comp.lower()
            if key in seen:
                continue
            seen.add(key)
            for m in range(1, 13):
                res[m] += sign * by_mes_orc[key].get(m, 0.0)
        return res
        
    real_row_vals: dict[int, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    orc_row_vals: dict[int, dict[int, float]] = defaultdict(lambda: defaultdict(float))
    
    resultado = []
    for r in rows:
        ordem = r.ordem
        tipo = r.tipo
        slug_str = r.agrupamento_slug
        formula = r.formula_texto
        
        vals_real = {}
        vals_orc = {}
        
        if tipo == "agrupamento":
            vals_real = get_val_real(slug_str)
            vals_orc = get_val_orc(slug_str)
            for m in range(1, 13):
                real_row_vals[ordem][m] = vals_real[m]
                orc_row_vals[ordem][m] = vals_orc[m]
        elif tipo == "totalizador":
            for m in range(1, 13):
                m_real_vals = {o: val_dict[m] for o, val_dict in real_row_vals.items()}
                m_orc_vals = {o: val_dict[m] for o, val_dict in orc_row_vals.items()}
                
                val_real = _eval_formula(formula, m_real_vals)
                val_orc = _eval_formula(formula, m_orc_vals)
                
                real_row_vals[ordem][m] = val_real
                orc_row_vals[ordem][m] = val_orc
                vals_real[m] = val_real
                vals_orc[m] = val_orc
                
        resultado.append({
            "ordem": ordem,
            "rotulo": r.rotulo,
            "tipo": tipo,
            "negrito_totalizador": bool(r.negrito_totalizador),
            "realizado": vals_real,
            "orcado": vals_orc,
        })
        
    return resultado


class FCOrcamentoUpsert(BaseModel):
    valor: float
    versao: str = "Original"


class IASugestaoRequest(BaseModel):
    agrupamento_slug: str
    rotulo: str
    cenario_usuario: str
    valores_referencia: List[float]


@router.get("/cliente/{cliente_id}/ano/{ano}/editavel")
def obter_orcamento_editavel(
    cliente_id: int,
    ano: int,
    versao: str = "Original",
    base: str = "fluxo_caixa",
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual)
):
    """Retorna a listagem de contas, totalizadores e títulos associados ao realizado do ano anterior e ao planejado atual."""
    # Check tenant restriction
    verificar_tenant(usuario, cliente_id)

    from routers.fc_exec import _parse_compound_slug, _eval_formula

    # 1. Carrega todas as linhas de template que sejam da base solicitada (fluxo_caixa ou dre)
    template_lines = db.execute(text("""
        SELECT tl.ordem, tl.rotulo, tl.tipo, tl.agrupamento_slug, tl.formula_texto, tl.negrito_totalizador
        FROM ref_template_linhas tl
        JOIN ref_templates t ON t.id = tl.template_id
        WHERE t.tipo = :base AND t.ativo = true
        ORDER BY tl.ordem
    """), {"base": base}).fetchall()

    if not template_lines:
        return []

    # 2. Busca valores de orçamento para o ano e versão especificados
    orc_raw = db.execute(text("""
        SELECT LOWER(agrupamento_slug) as slug, mes, valor
        FROM fc_orcamento
        WHERE cliente_id = :cid AND ano = :ano AND versao = :ver
    """), {"cid": cliente_id, "ano": ano, "ver": versao}).fetchall()

    by_mes_orc = defaultdict(lambda: defaultdict(float))
    for r in orc_raw:
        by_mes_orc[r.slug][r.mes] = float(r.valor)

    # 3. Busca valores realizados para o ano anterior (ano - 1)
    prev_ano = ano - 1
    real_raw = db.execute(text("""
        SELECT LOWER(agrupamento_slug) as slug, mes, COALESCE(SUM(valor), 0) as total
        FROM fc_lancamentos
        WHERE cliente_id = :cid AND ano = :prev_ano
        GROUP BY LOWER(agrupamento_slug), mes
    """), {"cid": cliente_id, "prev_ano": prev_ano}).fetchall()

    by_mes_real = defaultdict(lambda: defaultdict(float))
    for r in real_raw:
        by_mes_real[r.slug][r.mes] = float(r.total)

    # 4. Helpers para calcular valores agrupados simples ou compostos
    def calcular_valor_composto(slug_str, by_mes_map):
        if not slug_str:
            return {m: 0.0 for m in range(1, 13)}
        key_exact = slug_str.lower()
        if key_exact in by_mes_map:
            return {m: by_mes_map[key_exact].get(m, 0.0) for m in range(1, 13)}
            
        components = _parse_compound_slug(slug_str)
        result = {m: 0.0 for m in range(1, 13)}
        seen = set()
        for slug_comp, sign in components:
            key = slug_comp.lower()
            if key in seen:
                continue
            seen.add(key)
            for m in range(1, 13):
                result[m] += sign * by_mes_map[key].get(m, 0.0)
        return result

    # 5. Inicializa os vetores de valores de linha (agrupamentos vs totalizadores/títulos)
    row_vals_orc = {}
    row_vals_real = {}

    for line in template_lines:
        ordem = line.ordem
        if line.tipo == "agrupamento":
            row_vals_orc[ordem] = calcular_valor_composto(line.agrupamento_slug, by_mes_orc)
            row_vals_real[ordem] = calcular_valor_composto(line.agrupamento_slug, by_mes_real)
        else:
            row_vals_orc[ordem] = {m: 0.0 for m in range(1, 13)}
            row_vals_real[ordem] = {m: 0.0 for m in range(1, 13)}

    # 6. Resolve totalizadores em 10 passadas cruzando fórmulas de forward refs
    formula_map = {line.ordem: line.formula_texto for line in template_lines if line.tipo == "totalizador"}
    for _ in range(10):
        changed_orc = False
        changed_real = False
        for line in template_lines:
            if line.tipo != "totalizador":
                continue
            ordem = line.ordem
            formula = formula_map.get(ordem)
            if not formula:
                continue

            new_orc = {}
            new_real = {}
            for m in range(1, 13):
                scalars_orc = {k: v.get(m, 0.0) for k, v in row_vals_orc.items()}
                scalars_real = {k: v.get(m, 0.0) for k, v in row_vals_real.items()}
                new_orc[m] = _eval_formula(formula, scalars_orc)
                new_real[m] = _eval_formula(formula, scalars_real)

            if row_vals_orc[ordem] != new_orc:
                row_vals_orc[ordem] = new_orc
                changed_orc = True
            if row_vals_real[ordem] != new_real:
                row_vals_real[ordem] = new_real
                changed_real = True

        if not changed_orc and not changed_real:
            break

    # 7. Formata o payload de resposta em ordem
    resultado = []
    for line in template_lines:
        ordem = line.ordem
        slug = line.agrupamento_slug
        valores = {str(m): round(row_vals_orc[ordem].get(m, 0.0), 2) for m in range(1, 13)}
        realizado_ano_anterior = {str(m): round(row_vals_real[ordem].get(m, 0.0), 2) for m in range(1, 13)}

        resultado.append({
            "ordem": line.ordem,
            "rotulo": line.rotulo,
            "tipo": line.tipo or "agrupamento",
            "negrito_totalizador": bool(line.negrito_totalizador),
            "agrupamento_slug": slug,
            "formula_texto": line.formula_texto,
            "valores": valores,
            "realizado_ano_anterior": realizado_ano_anterior
        })

    return resultado


@router.put("/cliente/{cliente_id}/ano/{ano}/mes/{mes}/conta/{agrupamento_slug}")
def upsert_orcamento(
    cliente_id: int,
    ano: int,
    mes: int,
    agrupamento_slug: str,
    data: FCOrcamentoUpsert,
    db: Session = Depends(get_db),
    usuario=Depends(get_usuario_atual)
):
    """Insere ou atualiza um valor orçado mensal e gera registro de auditoria."""
    # Check tenant restriction
    verificar_tenant(usuario, cliente_id)

    if not (1 <= mes <= 12):
        raise HTTPException(status_code=400, detail="Mês inválido (deve ser entre 1 e 12)")

    # 1. Busca o registro de orçamento existente
    item = db.query(models.FCOrcamento).filter(
        models.FCOrcamento.cliente_id == cliente_id,
        models.FCOrcamento.agrupamento_slug == agrupamento_slug,
        models.FCOrcamento.ano == ano,
        models.FCOrcamento.mes == mes,
        models.FCOrcamento.versao == data.versao
    ).first()

    valor_antes = 0.0
    if item:
        valor_antes = item.valor
        item.valor = data.valor
    else:
        item = models.FCOrcamento(
            cliente_id=cliente_id,
            agrupamento_slug=agrupamento_slug,
            ano=ano,
            mes=mes,
            valor=data.valor,
            versao=data.versao
        )
        db.add(item)

    db.commit()

    # 2. Busca o rótulo legível do agrupamento para o log de auditoria
    rotulo = agrupamento_slug
    linha_template = db.execute(text("""
        SELECT tl.rotulo
        FROM ref_template_linhas tl
        JOIN ref_templates t ON t.id = tl.template_id
        WHERE t.tipo = 'fluxo_caixa' AND t.ativo = true AND tl.agrupamento_slug = :slug
        LIMIT 1
    """), {"slug": agrupamento_slug}).fetchone()
    if linha_template and linha_template.rotulo:
        rotulo = linha_template.rotulo

    # 3. Cria o log de auditoria usando helpers.log()
    from helpers import log as helpers_log
    helpers_log(
        db=db,
        usuario_id=usuario.id,
        acao="orcamento_editado",
        descricao=f"Orçamento de '{rotulo}' em {mes:02d}/{ano} (versão {data.versao}): de R$ {valor_antes:,.2f} para R$ {data.valor:,.2f}",
        projeto_id=None
    )

    return {"success": True}


@router.post("/cliente/{cliente_id}/ano/{ano}/sugerir-ia")
async def sugerir_ia_orcamento(
    cliente_id: int,
    ano: int,
    req: IASugestaoRequest,
    usuario=Depends(get_usuario_atual)
):
    """Chama a IA para projetar a série orçamentária baseado no realizado anterior e no cenário."""
    # Check tenant restriction
    verificar_tenant(usuario, cliente_id)

    api_key_gemini = os.getenv("GEMINI_API_KEY")
    api_key_claude = os.getenv("ANTHROPIC_API_KEY")
    api_key_openrouter = os.getenv("OPENROUTER_API_KEY")

    if not (api_key_gemini or api_key_claude or api_key_openrouter):
        raise HTTPException(status_code=503, detail="Nenhum serviço de IA configurado no ambiente.")

    system_prompt = (
        "Você é um copiloto de planejamento orçamentário. "
        "Sua tarefa é projetar ou ajustar uma série mensal de orçamento (12 meses, de Janeiro a Dezembro) "
        "baseando-se no histórico de realizado do ano anterior e no cenário/premissas descritos pelo consultor. "
        "Responda APENAS com um array JSON contendo 12 números de ponto flutuante, correspondentes aos meses de Janeiro a Dezembro. "
        "Não inclua nenhuma outra palavra, explicação, formatação em Markdown (como ```json) ou texto no resultado."
    )

    user_text = (
        f"Conta/Categoria: {req.rotulo} (slug: {req.agrupamento_slug})\n"
        f"Realizado do Ano Anterior (Jan a Dez): {req.valores_referencia}\n"
        f"Cenário/Premissa do Consultor: {req.cenario_usuario}\n\n"
        "Projete os 12 meses de orçamento seguindo estritamente as regras e retorne apenas o array JSON."
    )

    resposta_texto = ""
    modelo_usado = ""

    # Tenta Gemini primeiro
    if api_key_gemini:
        gemini_model = os.getenv("GEMINI_MODEL", "gemini-2.0-flash")
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{gemini_model}:generateContent?key={api_key_gemini}"
        payload = {
            "contents": [{"role": "user", "parts": [{"text": user_text}]}],
            "systemInstruction": {"parts": [{"text": system_prompt}]},
            "generationConfig": {"maxOutputTokens": 256, "temperature": 0.3},
        }
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.post(url, json=payload)
                if resp.status_code == 200:
                    candidates = resp.json().get("candidates", [])
                    if candidates:
                        resposta_texto = candidates[0]["content"]["parts"][0]["text"]
                        modelo_usado = gemini_model
        except Exception:
            pass

    # Tenta Claude se Gemini falhou ou não estava configurado
    if not resposta_texto and api_key_claude:
        claude_model = os.getenv("ANTHROPIC_MODEL", "claude-haiku-4-5-20251001")
        payload = {
            "model": claude_model,
            "max_tokens": 256,
            "system": system_prompt,
            "messages": [{"role": "user", "content": user_text}],
            "temperature": 0.3
        }
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key": api_key_claude,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json",
                    },
                    json=payload
                )
                if resp.status_code == 200:
                    resposta_texto = resp.json()["content"][0]["text"]
                    modelo_usado = claude_model
        except Exception:
            pass

    # Tenta OpenRouter como última opção
    if not resposta_texto and api_key_openrouter:
        payload = {
            "model": "google/gemini-2.0-flash-001",
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_text},
            ],
            "max_tokens": 256,
            "temperature": 0.3,
        }
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.post(
                    "https://openrouter.ai/api/v1/chat/completions",
                    json=payload,
                    headers={
                        "Authorization": f"Bearer {api_key_openrouter}",
                        "Content-Type": "application/json"
                    }
                )
                if resp.status_code == 200:
                    choices = resp.json().get("choices", [])
                    if choices:
                        resposta_texto = choices[0]["message"]["content"]
                        modelo_usado = "openrouter/gemini-2.0-flash"
        except Exception:
            pass

    if not resposta_texto:
        raise HTTPException(status_code=502, detail="Nenhum provedor de IA respondeu com sucesso.")

    # Parse da resposta (limpando qualquer formatação markdown adicional)
    clean_text = resposta_texto.strip()
    clean_text = re.sub(r"^```(?:json)?\s*", "", clean_text)
    clean_text = re.sub(r"\s*```$", "", clean_text)
    clean_text = clean_text.strip()

    try:
        valores = json.loads(clean_text)
        if isinstance(valores, list) and len(valores) == 12:
            return {
                "success": True,
                "valores": [float(x) for x in valores],
                "modelo": modelo_usado
            }
        else:
            raise ValueError("O formato retornado não é um array com 12 números")
    except Exception as e:
        raise HTTPException(
            status_code=502,
            detail=f"Erro ao interpretar resposta da IA como série mensal: {str(e)}. Resposta bruta: {resposta_texto}"
        )

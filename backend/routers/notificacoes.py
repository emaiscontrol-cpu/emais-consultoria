from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from datetime import datetime, timezone
from io import BytesIO
from database import get_db
from security import get_usuario_atual
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import models

router = APIRouter()

TIPO_PT = {
    "tarefa_atrasada": "Tarefa atrasada",
    "prazo_proximo":   "Prazo se aproximando",
}


def gerar_alertas(db: Session, usuario: models.Usuario):
    agora = datetime.now(timezone.utc)
    alertas = []

    q = db.query(models.Projeto).options(
        joinedload(models.Projeto.fases).joinedload(models.Fase.tarefas)
    )
    if usuario.perfil in ("analista", "ger_projeto", "ti"):
        q = q.filter(models.Projeto.cliente_id == usuario.cliente_id)

    for projeto in q.all():
        for fase in projeto.fases:
            for tarefa in fase.tarefas:
                if not tarefa.data_prazo or tarefa.status.value == "concluida":
                    continue
                prazo = tarefa.data_prazo
                if prazo.tzinfo is None:
                    prazo = prazo.replace(tzinfo=timezone.utc)
                diff = (prazo - agora).days
                if diff < 0:
                    alertas.append({
                        "tipo": "tarefa_atrasada",
                        "titulo": f"Atrasada {abs(diff)}d: {tarefa.nome}",
                        "mensagem": f"{projeto.nome} › {fase.nome}",
                        "projeto_id": projeto.id,
                        "projeto_nome": projeto.nome,
                        "tarefa_id": tarefa.id,
                        "dias": abs(diff),
                    })
                elif diff <= 3:
                    alertas.append({
                        "tipo": "prazo_proximo",
                        "titulo": f"Vence em {diff}d: {tarefa.nome}",
                        "mensagem": f"{projeto.nome} › {fase.nome}",
                        "projeto_id": projeto.id,
                        "projeto_nome": projeto.nome,
                        "tarefa_id": tarefa.id,
                        "dias": diff,
                    })

    alertas.sort(key=lambda a: a["dias"])
    return alertas


@router.get("/")
def listar(db: Session = Depends(get_db), usuario=Depends(get_usuario_atual)):
    alertas = gerar_alertas(db, usuario)
    mencoes = (
        db.query(models.NotificacaoMencao)
        .filter(models.NotificacaoMencao.usuario_destino_id == usuario.id)
        .order_by(models.NotificacaoMencao.criado_em.desc())
        .limit(30)
        .all()
    )
    for m in mencoes:
        alertas.append({
            "tipo": "mencao",
            "id_mencao": m.id,
            "lida": m.lida,
            "titulo": m.mensagem,
            "mensagem": f"de {m.de_usuario.nome}",
            "projeto_id": m.projeto_id,
            "dias": 0,
        })
    return alertas


@router.post("/mencao/{id}/lida", status_code=204)
def marcar_lida(id: int, db: Session = Depends(get_db), usuario=Depends(get_usuario_atual)):
    m = db.query(models.NotificacaoMencao).filter(
        models.NotificacaoMencao.id == id,
        models.NotificacaoMencao.usuario_destino_id == usuario.id,
    ).first()
    if m:
        m.lida = True
        db.commit()


@router.get("/excel")
def relatorio_excel(db: Session = Depends(get_db), usuario=Depends(get_usuario_atual)):
    alertas = gerar_alertas(db, usuario)

    wb = Workbook()
    ws = wb.active
    ws.title = "Notificações"

    headers = ["Tipo", "Título", "Detalhe", "Projeto", "Dias"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="1E40AF")
        c.font = Font(bold=True, color="FFFFFF")

    for row, a in enumerate(alertas, 2):
        ws.cell(row=row, column=1, value=TIPO_PT.get(a["tipo"], a["tipo"]))
        ws.cell(row=row, column=2, value=a["titulo"])
        ws.cell(row=row, column=3, value=a["mensagem"])
        ws.cell(row=row, column=4, value=a["projeto_nome"])
        ws.cell(row=row, column=5, value=a["dias"])

    for col in range(1, 6):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 35

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome = f"notificacoes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"},
    )

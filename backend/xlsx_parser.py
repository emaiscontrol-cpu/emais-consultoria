"""
Parser de XLSX de realizado (ERP) usando openpyxl.

Suporta dois tipos de estrutura configurados em ImportLayout:
  COLUNAS_MESES  → cada mês é uma coluna (formato tabular mais comum)
  LINHA_MES_VALOR → cada linha tem (conta, mês, valor) — formato diário/lançamento
"""

import io
import json
from models import ImportLayout

MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def _val(cell) -> float:
    if cell is None:
        return 0.0
    s = str(cell).strip().replace(",", ".").replace(" ", "").replace("R$", "").replace(".", "", str(cell).count(".") - 1)
    # segunda passagem: remover separador de milhar residual
    try:
        return float(s)
    except Exception:
        return 0.0


def _limpar_val(cell) -> float:
    """Converte célula para float, tratando vírgula decimal e separadores de milhar."""
    if cell is None:
        return 0.0
    s = str(cell).strip()
    # Remove R$, espaços e símbolos de moeda
    for ch in ("R$", "$", "€", " "):
        s = s.replace(ch, "")
    # Detectar separador decimal: se vírgula aparece como último separador → decimal BR
    if "," in s and "." in s:
        # Formato 1.234,56 → remover pontos e trocar vírgula
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _str(cell) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def _parse_mes(val, formato: str) -> int | None:
    if val is None:
        return None
    s = _str(val).lower()
    if s in MESES_PT:
        return MESES_PT[s]
    if "/" in s:
        parts = s.split("/")
        try:
            m = int(parts[0])
            if 1 <= m <= 12:
                return m
        except Exception:
            pass
    if "-" in s:
        parts = s.split("-")
        try:
            m = int(parts[1] if len(parts) > 1 else parts[0])
            if 1 <= m <= 12:
                return m
        except Exception:
            pass
    try:
        m = int(s)
        if 1 <= m <= 12:
            return m
    except Exception:
        pass
    return None


def parse_xlsx(content: bytes, layout: ImportLayout) -> list[dict]:
    """
    Lê o XLSX e retorna lista de:
      {"codigo": str, "descricao": str, "mes": int | None, "valor": float, "unidade": str | None}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    prefixos = json.loads(layout.prefixos_ignorar or "[]")
    linhas_ignorar = set(json.loads(layout.linhas_ignorar or "[]"))
    col_conta = layout.coluna_conta  # 0-based
    col_desc = layout.coluna_descricao

    resultado = []

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if layout.tipo_estrutura == "COLUNAS_MESES":
        mapa = json.loads(layout.mapa_colunas_meses or "[]")
        col_uni = layout.coluna_unidade

        for row_num, row in enumerate(rows, start=1):
            if row_num < layout.linha_inicio:
                continue
            if row_num in linhas_ignorar:
                continue
            if not any(c for c in row if c is not None):
                continue

            codigo = _str(row[col_conta] if col_conta < len(row) else None)
            if not codigo:
                continue
            if any(codigo.startswith(p) for p in prefixos if p):
                continue

            desc = _str(row[col_desc] if col_desc is not None and col_desc < len(row) else None)
            unidade = _str(row[col_uni]) if col_uni is not None and col_uni < len(row) else None

            for entry in mapa:
                mes = int(entry["mes"])
                col = int(entry["coluna"])
                valor = _limpar_val(row[col] if col < len(row) else None)
                if valor != 0.0:
                    resultado.append({
                        "codigo": codigo,
                        "descricao": desc,
                        "mes": mes,
                        "valor": valor,
                        "unidade": unidade,
                    })

    elif layout.tipo_estrutura == "LINHA_MES_VALOR":
        col_mes = layout.coluna_mes or 1
        col_val = layout.coluna_valor or 2
        col_uni = layout.coluna_unidade

        for row_num, row in enumerate(rows, start=1):
            if row_num < layout.linha_inicio:
                continue
            if row_num in linhas_ignorar:
                continue
            if not any(c for c in row if c is not None):
                continue

            codigo = _str(row[col_conta] if col_conta < len(row) else None)
            if not codigo:
                continue
            if any(codigo.startswith(p) for p in prefixos if p):
                continue

            desc = _str(row[col_desc] if col_desc is not None and col_desc < len(row) else None)
            mes = _parse_mes(row[col_mes] if col_mes < len(row) else None, layout.formato_mes or "MM/YYYY")
            if not mes:
                continue
            valor = _limpar_val(row[col_val] if col_val < len(row) else None)
            unidade = _str(row[col_uni]) if col_uni is not None and col_uni < len(row) else None

            resultado.append({
                "codigo": codigo,
                "descricao": desc,
                "mes": mes,
                "valor": valor,
                "unidade": unidade,
            })

    elif layout.tipo_estrutura == "COLUNAS_UNIDADES":
        # Estrutura com colunas de unidades/filiais lado a lado (um único mês importado)
        linha_cab_idx = max(0, layout.linha_inicio - 2)
        if len(rows) > linha_cab_idx:
            cabecalho = rows[linha_cab_idx]
        else:
            cabecalho = []

        col_inicio_uni = layout.coluna_inicio_unidades or 1

        for row_num, row in enumerate(rows, start=1):
            if row_num < layout.linha_inicio:
                continue
            if row_num in linhas_ignorar:
                continue
            if not any(c for c in row if c is not None):
                continue

            codigo = _str(row[col_conta] if col_conta < len(row) else None)
            if not codigo:
                continue
            if any(codigo.startswith(p) for p in prefixos if p):
                continue

            desc = _str(row[col_desc] if col_desc is not None and col_desc < len(row) else None)

            # Lança uma entrada por coluna de unidade a partir de col_inicio_uni
            for col_idx in range(col_inicio_uni, len(row)):
                # Nome da unidade vem do cabeçalho correspondente
                nome_unidade = _str(cabecalho[col_idx]) if col_idx < len(cabecalho) else f"Unidade_{col_idx}"
                if not nome_unidade or nome_unidade.lower() in ("consolidado", "total", "diferença", "diferenca", "geral", "desvio", "%", "part."):
                    continue

                valor = _limpar_val(row[col_idx])
                if valor != 0.0:
                    resultado.append({
                        "codigo": codigo,
                        "descricao": desc,
                        "mes": None, # Preenchido no endpoint com o mês selecionado
                        "valor": valor,
                        "unidade": nome_unidade
                    })

    return resultado


def preview_xlsx(content: bytes, n_rows: int = 5) -> dict:
    """Retorna cabeçalho e primeiras linhas para configuração do layout."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append([str(c) if c is not None else "" for c in row])
        if i >= n_rows:
            break
    wb.close()
    return {"linhas": rows, "n_colunas": max((len(r) for r in rows), default=0)}
